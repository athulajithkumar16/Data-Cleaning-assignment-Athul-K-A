import pdfplumber
import pandas as pd
import os
import re
import warnings
from datetime import datetime
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

def extract_pdf_data(pdf_path):
    try:
        with pdfplumber.open(pdf_path) as pdf:
            page = pdf.pages[0]
            
            text = page.extract_text(x_tolerance=3, y_tolerance=3)
            if not text:
                text = page.extract_text_simple(x_tolerance=5)
            
            if not text:
                print(f"Warning: No text extracted from {pdf_path}")
                return None

            brand_patterns = [
                r"Restaurant / Store ID\s*:\s*(\d+)",
                r"Store ID\s*:\s*(\d+)",
                r"ID\s*:\s*(\d+)"
            ]
            
            brand_id = None
            for pattern in brand_patterns:
                match = re.search(pattern, text)
                if match:
                    brand_id = match.group(1)
                    break

            amount_patterns = [
                r"Grand Total\s*:\s*([\d,]+\.\d{2})",
                r"Grand Total\s*Rs\.\s*([\d,]+\.\d{2})",
                r"Total Amount \(Rs\.\)\s*([\d,]+\.\d{2})",
                r"Grand Total[\s\S]*?(\d[\d,]*\.\d{2})",
                r"Total\s*Amount\s*\(Rs\.\)[\s\S]*?(\d[\d,]*\.\d{2})",
                r"Amount\s*in\s*Words[\s\S]*?(\d[\d,]*\.\d{2})"
            ]
            
            grand_total = None
            for pattern in amount_patterns:
                match = re.search(pattern, text.replace("\n", " "))
                if match:
                    try:
                        grand_total = float(match.group(1).replace(",", ""))
                        break
                    except ValueError:
                        continue
            
            if grand_total is None:
                print(f"Warning: Could not find Grand Total in {pdf_path}")
                return None
            
            return {
                "payout_period": "01/04/2025 to 05/04/2025",
                "file_name": os.path.basename(pdf_path),
                "brand_id": brand_id or "N/A",
                "description": "Service Fee",
                "base_amount": round(grand_total / 1.18, 2),
                "grand_total": grand_total,
                "extraction_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
    except Exception as e:
        print(f"Error processing {pdf_path}: {str(e)}")
        return None

def main():
    print("Starting PDF processing...")
    
    pdf_files = [f for f in os.listdir() if f.lower().endswith(".pdf")]
    pdf_data = []
    
    for pdf in pdf_files:
        print(f"Processing {pdf}...")
        data = extract_pdf_data(pdf)
        if data:
            pdf_data.append(data)
    
    if not pdf_data:
        print("Error: No PDFs processed successfully")
        return
    
    pdf_df = pd.DataFrame(pdf_data)
    output_file = "Output_Commission_Invoice.xlsx"
    
    try:
        wb = load_workbook("Swiggy_Tax_Sample_file.xlsx")
    except FileNotFoundError:
        wb = load_workbook()
    
    if "Commission Invoice" not in wb.sheetnames:
        wb.create_sheet("Commission Invoice")
    
    ws = wb["Commission Invoice"]
    
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    
    for row in pdf_df.itertuples():
        ws.append([
            row.payout_period, row.file_name, "FY_2024-2025", "2025", "April", 
            "", "29ABNFM9601R1Z9", "29AAFCB7707D1ZQ", 1, row.description, 
            "996211", "OTH", 1, row.base_amount, 0, row.base_amount, 
            9, round(row.base_amount * 0.09, 2), 9, round(row.base_amount * 0.09, 2), 
            0, 0, 0, 0, 0, 0, row.grand_total, 
            0, row.grand_total, row.brand_id, "AAFCB7707D", 
            "2025-04-09", row.file_name.split("_")[-1].replace(".pdf", ""), "", "INV", "",
            row.extraction_time
        ])
    
    wb.save(output_file)
    print(f"\nSuccess! Processed {len(pdf_data)}/{len(pdf_files)} PDFs")
    print(f"Output saved to {output_file}")
    
    failed = set(pdf_files) - set([os.path.basename(x['file_name']) for x in pdf_data])
    if failed:
        print("\nFailed to process:")
        for f in failed:
            print(f"- {f}")

if __name__ == "__main__":
    main()